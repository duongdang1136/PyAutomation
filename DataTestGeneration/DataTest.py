from faker import Faker
from openpyxl import workbook
from openpyxl.workbook import Workbook

wb = Workbook()
ws = wb.active
fake_data = Faker('en_US')

for i in range(1, 200):
    for j in range(1, 4):
        ws.cell(row=i, column=1).value = fake_data.name()
        ws.cell(row=i, column=2).value = fake_data.email()
        ws.cell(row=i, column=3).value = fake_data.address()
wb.save("testData.xlsx")